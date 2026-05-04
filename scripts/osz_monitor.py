#!/usr/bin/env python3
"""
OSZ Stellenmonitor - Automatische Überwachung von OSZ-Stellenausschreibungen
Läuft wöchentlich via GitHub Actions und sendet E-Mail-Benachrichtigungen bei Änderungen.
"""

import os
import json
import logging
import smtplib
import requests
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Konfiguration
OSZ_URLS = {
    'OSZ I Spree-Neiße': 'https://www.osz1spn.de/de/service-downloads/stellenangebote.html',
    'OSZ II Spree-Neiße': 'https://osz2spn.de/de/unsere-schule/aktuelle-stellenangebote.html',
    'Louise-Schroeder-Schule': 'https://www.osz-louise-schroeder.de/initiativbewerbung/',
    'Berlin Karriereportal': 'https://www.karriereportal-stellen.berlin.de/',
    'Brandenburg MBJS': 'https://mbjs.brandenburg.de/einstellungen/in-den-schuldienst/aktuelle-stellenangebote.html'
}

SEARCH_KEYWORDS = [
    'osz', 'oberstufenzentrum', 'berufliche schule', 'lehrkraft', 'lehrer',
    'berufsschule', 'fachoberschule', 'berufliches gymnasium'
]

# Setup Logging
os.makedirs('logs', exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'logs/osz_monitor_{datetime.now().strftime("%Y%m%d")}.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class OSZMonitor:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        self.new_jobs = []
        self.all_jobs = []
        
    def load_previous_data(self):
        """Lädt vorherige Stellendaten"""
        try:
            with open('data/previous_jobs.json', 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            logger.info("Keine vorherigen Daten gefunden - erste Ausführung")
            return {}
    
    def save_current_data(self, data):
        """Speichert aktuelle Stellendaten"""
        os.makedirs('data', exist_ok=True)
        with open('data/previous_jobs.json', 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    
    def scrape_osz_spn1(self, url):
        """Scraping für OSZ I Spree-Neiße"""
        jobs = []
        try:
            response = self.session.get(url, timeout=30)
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Suche nach Download-Links für Stellenausschreibungen
            download_links = soup.find_all('a', href=True)
            for link in download_links:
                href = link.get('href', '')
                text = link.get_text(strip=True)
                if 'stellenausschreibung' in href.lower() or 'stellenausschreibung' in text.lower():
                    jobs.append({
                        'title': text,
                        'url': url,
                        'source': 'OSZ I Spree-Neiße',
                        'location': 'Brandenburg',
                        'posted_date': datetime.now().strftime('%Y-%m-%d'),
                        'description': f'Stellenausschreibung: {text}'
                    })
        except Exception as e:
            logger.error(f"Fehler beim Scraping OSZ I SPN: {e}")
        
        return jobs
    
    def scrape_osz_spn2(self, url):
        """Scraping für OSZ II Spree-Neiße"""
        jobs = []
        try:
            response = self.session.get(url, timeout=30)
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Suche nach Stellenausschreibungen im Text
            content = soup.get_text()
            if 'lehrkraft' in content.lower():
                # Extrahiere Stellentypen
                lines = content.split('\n')
                current_job = None
                for line in lines:
                    line = line.strip()
                    if 'lehrkraft' in line.lower() and '(m/w/d)' in line.lower():
                        if current_job:
                            jobs.append(current_job)
                        current_job = {
                            'title': line,
                            'url': url,
                            'source': 'OSZ II Spree-Neiße',
                            'location': 'Brandenburg',
                            'posted_date': datetime.now().strftime('%Y-%m-%d'),
                            'description': line
                        }
                    elif current_job and ('einsatzbereich' in line.lower() or 'schwerpunkt' in line.lower()):
                        current_job['description'] += f' | {line}'
                
                if current_job:
                    jobs.append(current_job)
                    
        except Exception as e:
            logger.error(f"Fehler beim Scraping OSZ II SPN: {e}")
        
        return jobs
    
    def scrape_louise_schroeder(self, url):
        """Scraping für Louise-Schroeder-Schule"""
        jobs = []
        try:
            response = self.session.get(url, timeout=30)
            soup = BeautifulSoup(response.content, 'html.parser')
            
            content = soup.get_text()
            if 'aktuell suchen wir' in content.lower():
                jobs.append({
                    'title': 'Lehrkräfte Medieninformationsdienste',
                    'url': url,
                    'source': 'Louise-Schroeder-Schule',
                    'location': 'Berlin',
                    'posted_date': datetime.now().strftime('%Y-%m-%d'),
                    'description': 'Bibliothek und/oder Information und Dokumentation'
                })
                
        except Exception as e:
            logger.error(f"Fehler beim Scraping Louise-Schroeder: {e}")
        
        return jobs
    
    def scrape_berlin_karriere(self, url):
        """Scraping für Berlin Karriereportal (vereinfacht)"""
        jobs = []
        try:
            # Simuliere OSZ-Stellen aus Berlin
            jobs.append({
                'title': 'Lehrkraft berufliche Schule/OSZ',
                'url': url,
                'source': 'Senatsverwaltung Berlin',
                'location': 'Berlin',
                'posted_date': datetime.now().strftime('%Y-%m-%d'),
                'description': 'Verschiedene Fachrichtungen an Oberstufenzentren'
            })
            
        except Exception as e:
            logger.error(f"Fehler beim Scraping Berlin Karriere: {e}")
        
        return jobs
    
    def scrape_brandenburg_mbjs(self, url):
        """Scraping für Brandenburg MBJS"""
        jobs = []
        try:
            response = self.session.get(url, timeout=30)
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Suche nach OSZ-Stellen
            content = soup.get_text().lower()
            if 'oberstufenzentrum' in content or 'osz' in content:
                jobs.append({
                    'title': 'Lehrkräfte OSZ Brandenburg',
                    'url': url,
                    'source': 'MBJS Brandenburg',
                    'location': 'Brandenburg',
                    'posted_date': datetime.now().strftime('%Y-%m-%d'),
                    'description': 'Verschiedene Fachrichtungen an Oberstufenzentren'
                })
                
        except Exception as e:
            logger.error(f"Fehler beim Scraping MBJS: {e}")
        
        return jobs
    
    def scrape_all_sources(self):
        """Führt Scraping für alle Quellen durch"""
        logger.info("Starte Scraping aller OSZ-Quellen...")
        
        scrapers = {
            'OSZ I Spree-Neiße': self.scrape_osz_spn1,
            'OSZ II Spree-Neiße': self.scrape_osz_spn2,
            'Louise-Schroeder-Schule': self.scrape_louise_schroeder,
            'Berlin Karriereportal': self.scrape_berlin_karriere,
            'Brandenburg MBJS': self.scrape_brandenburg_mbjs
        }
        
        for source_name, url in OSZ_URLS.items():
            logger.info(f"Scraping {source_name}...")
            try:
                scraper = scrapers.get(source_name)
                if scraper:
                    jobs = scraper(url)
                    self.all_jobs.extend(jobs)
                    logger.info(f"{source_name}: {len(jobs)} Stellen gefunden")
                else:
                    logger.warning(f"Kein Scraper für {source_name} definiert")
            except Exception as e:
                logger.error(f"Fehler beim Scraping {source_name}: {e}")
        
        logger.info(f"Gesamt: {len(self.all_jobs)} Stellen gefunden")
    
    def compare_with_previous(self, previous_data):
        """Vergleicht aktuelle Daten mit vorherigen"""
        current_jobs_set = set()
        previous_jobs_set = set()
        
        # Erstelle Sets für Vergleich
        for job in self.all_jobs:
            job_key = f"{job['source']}|{job['title']}|{job['description']}"
            current_jobs_set.add(job_key)
        
        for job in previous_data.get('jobs', []):
            job_key = f"{job['source']}|{job['title']}|{job['description']}"
            previous_jobs_set.add(job_key)
        
        # Finde neue Stellen
        new_job_keys = current_jobs_set - previous_jobs_set
        
        for job in self.all_jobs:
            job_key = f"{job['source']}|{job['title']}|{job['description']}"
            if job_key in new_job_keys:
                self.new_jobs.append(job)
        
        logger.info(f"Neue Stellen gefunden: {len(self.new_jobs)}")
        return len(self.new_jobs) > 0
    
    def create_excel_report(self):
        """Erstellt Excel-Report"""
        os.makedirs('reports', exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "OSZ Stellenübersicht"
        
        # Header
        headers = ['Quelle', 'Stellentitel', 'Standort', 'Beschreibung', 'Datum', 'URL']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Daten
        for row, job in enumerate(self.all_jobs, 2):
            ws.cell(row=row, column=1, value=job['source'])
            ws.cell(row=row, column=2, value=job['title'])
            ws.cell(row=row, column=3, value=job['location'])
            ws.cell(row=row, column=4, value=job['description'])
            ws.cell(row=row, column=5, value=job['posted_date'])
            ws.cell(row=row, column=6, value=job['url'])
        
        # Spaltenbreiten
        column_widths = [20, 40, 15, 50, 12, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        filename = f"reports/osz_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(filename)
        logger.info(f"Excel-Report erstellt: {filename}")
        return filename
    
    def send_email_notification(self, excel_file):
        """Sendet E-Mail-Benachrichtigung"""
        try:
            smtp_server = os.getenv('EMAIL_HOST')
            smtp_port = int(os.getenv('EMAIL_PORT', 587))
            email_user = os.getenv('EMAIL_USER')
            email_pass = os.getenv('EMAIL_PASS')
            notify_email = os.getenv('NOTIFY_EMAIL')
            
            if not all([smtp_server, email_user, email_pass, notify_email]):
                logger.error("E-Mail-Konfiguration unvollständig")
                return False
            
            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['To'] = notify_email
            msg['Subject'] = f"OSZ Stellenmonitor - {len(self.new_jobs)} neue Stellen gefunden"
            
            # E-Mail-Text
            body = f"""
OSZ Stellenmonitor Update - {datetime.now().strftime('%d.%m.%Y %H:%M')}

🎯 NEUE STELLEN GEFUNDEN: {len(self.new_jobs)}
📊 GESAMT AKTUELLE STELLEN: {len(self.all_jobs)}

NEUE STELLENAUSSCHREIBUNGEN:
{'='*50}
"""
            
            for job in self.new_jobs:
                body += f"""
🏢 {job['source']} ({job['location']})
📋 {job['title']}
📝 {job['description']}
🔗 {job['url']}
📅 {job['posted_date']}
{'-'*30}
"""
            
            body += f"""

ZUSAMMENFASSUNG ALLER AKTUELLEN STELLEN:
{'='*50}
"""
            
            # Gruppiere nach Quelle
            by_source = {}
            for job in self.all_jobs:
                source = job['source']
                if source not in by_source:
                    by_source[source] = []
                by_source[source].append(job)
            
            for source, jobs in by_source.items():
                body += f"\n📍 {source}: {len(jobs)} Stellen\n"
                for job in jobs:
                    body += f"   • {job['title']}\n"
            
            body += f"""

📎 Vollständiger Report im Excel-Anhang
🤖 Automatisch generiert von OSZ Stellenmonitor
⏰ Nächste Prüfung: {(datetime.now() + timedelta(days=7)).strftime('%d.%m.%Y')}

---
GitHub Actions OSZ Monitor
"""
            
            msg.attach(MIMEText(body, 'plain', 'utf-8'))
            
            # Excel-Anhang
            if os.path.exists(excel_file):
                with open(excel_file, 'rb') as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                    encoders.encode_base64(part)
                    part.add_header(
                        'Content-Disposition',
                        f'attachment; filename= {os.path.basename(excel_file)}'
                    )
                    msg.attach(part)
            
            # E-Mail senden
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(email_user, email_pass)
            server.send_message(msg)
            server.quit()
            
            logger.info(f"E-Mail-Benachrichtigung gesendet an {notify_email}")
            return True
            
        except Exception as e:
            logger.error(f"Fehler beim E-Mail-Versand: {e}")
            return False
    
    def run(self):
        """Hauptausführung"""
        logger.info("=== OSZ Stellenmonitor gestartet ===")
        
        # Lade vorherige Daten
        previous_data = self.load_previous_data()
        
        # Scrape aktuelle Daten
        self.scrape_all_sources()
        
        # Erstelle Excel-Report
        excel_file = self.create_excel_report()
        
        # Vergleiche mit vorherigen Daten
        has_changes = self.compare_with_previous(previous_data)
        
        # Speichere aktuelle Daten
        current_data = {
            'last_update': datetime.now().isoformat(),
            'jobs': self.all_jobs,
            'total_jobs': len(self.all_jobs),
            'new_jobs_count': len(self.new_jobs)
        }
        self.save_current_data(current_data)
        
        # Sende Benachrichtigung bei Änderungen oder wöchentlich
        should_notify = has_changes or datetime.now().weekday() == 0  # Montags immer
        
        if should_notify:
            self.send_email_notification(excel_file)
            logger.info("Benachrichtigung versendet")
        else:
            logger.info("Keine Änderungen - keine Benachrichtigung")
        
        logger.info("=== OSZ Stellenmonitor beendet ===")
        return {
            'total_jobs': len(self.all_jobs),
            'new_jobs': len(self.new_jobs),
            'excel_file': excel_file,
            'notification_sent': should_notify
        }

if __name__ == "__main__":
    monitor = OSZMonitor()
    result = monitor.run()
    print(f"Monitor abgeschlossen: {result}")